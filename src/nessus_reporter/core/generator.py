# src/nessus_reporter/core/generator.py

import pandas as pd
import logging
from pathlib import Path
from typing import List

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ReportGenerationError(Exception):
    """當生成報告過程中發生錯誤時引發的基礎類別。"""
    pass

class ExcelReportGenerator:
    """
    負責將 pandas DataFrame 生成為格式精美的 Excel 檔案。
    這個類別的所有方法均為靜態方法，因為它不儲存任何狀態。
    """
    # [優化] 將樣式值定義為常數，方便統一管理
    SHEET_NAME = "Nessus Report"
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")
    HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    FREEZE_PANE_CELL = "A2"

    @staticmethod
    def _apply_styles(worksheet: Worksheet) -> None:
        """私有輔助方法：統一應用所有靜態格式（標頭、凍結窗格、篩選器）。"""
        # 1. 格式化標頭
        for cell in worksheet[1]:
            cell.font = ExcelReportGenerator.HEADER_FONT
            cell.fill = ExcelReportGenerator.HEADER_FILL
            cell.alignment = ExcelReportGenerator.HEADER_ALIGNMENT
        
        # 2. 凍結首行
        worksheet.freeze_panes = ExcelReportGenerator.FREEZE_PANE_CELL
        
        # 3. 增加自動篩選器
        worksheet.auto_filter.ref = worksheet.dimensions

    @staticmethod
    def generate_report(
        df: pd.DataFrame, 
        selected_columns: List[str], 
        output_path: Path
    ) -> None:
        """
        接收 DataFrame，篩選指定欄位，並生成一個格式化的 Excel 報告。
        """
        if df.empty:
            logging.info("傳入的 DataFrame 為空，已跳過生成報告。")
            return
            
        final_columns = [col for col in selected_columns if col in df.columns]
        if not final_columns:
            logging.warning("沒有有效的欄位被選取，已跳過生成報告。")
            return
            
        report_df = df[final_columns]

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                report_df.to_excel(writer, sheet_name=ExcelReportGenerator.SHEET_NAME, index=False)
                worksheet = writer.sheets[ExcelReportGenerator.SHEET_NAME]

                # [優化] 使用更高效能的方式計算並設定欄寬
                for i, col_name in enumerate(report_df.columns, 1):
                    # 計算標頭和內容的最大長度
                    # .astype(str).map(len) 是高效的向量化操作
                    max_len_content = report_df[col_name].astype(str).map(len).max()
                    max_len_header = len(col_name)
                    max_len = max(max_len_content, max_len_header)
                    
                    # 設定欄寬
                    adjusted_width = (max_len + 2)
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                
                # 套用其他靜態格式
                ExcelReportGenerator._apply_styles(worksheet)

        except PermissionError:
            raise ReportGenerationError(f"無法寫入檔案，請確認 '{output_path.name}' 沒有被其他程式打開。")
        except Exception as e:
            raise ReportGenerationError(f"生成報告時發生未預期的錯誤: {e}") from e
